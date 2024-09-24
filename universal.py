import pandas as pd
import tkinter as tk
from tkinter import filedialog
import shutil
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor
import time
import os

def process_excel_file():
    start_time = time.time()  # Start timing

    # Открытие диалогового окна для выбора исходного файла
    root = tk.Tk()
    root.withdraw()
    input_file_path = filedialog.askopenfilename(title="Выберите исходный Excel файл", filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    if not input_file_path:
        print("Файл не выбран.")
        return
    
    # Создание копии выбранного файла
    output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Сохраните обработанный файл", filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not output_file_path:
        print("Местоположение для сохранения файла не выбрано.")
        return
    
    shutil.copy(input_file_path, output_file_path)
    
    # Чтение данных из Excel файла
    try:
        df = pd.read_excel(output_file_path, sheet_name=None)
        df = pd.concat(df.values(), ignore_index=True)  # Объединение всех листов, если их несколько
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return
    
    # Сохранение неизменённых данных (индексы 0-11)
    unchanged_data = df.iloc[:, :11].copy()
    
    # Обработка данных с индекса 11 и далее
    data = df.iloc[:, 11:]
    
    # Определение границ для разделения данных
    num_columns = len(data.columns)
    letter_index = next((i for i, col in enumerate(data.columns) if any(char.isalpha() for char in str(col))), num_columns)
    
    # Разделение данных на три части
    third = letter_index // 3
    abc_answers = data.iloc[:, :third].copy()
    correct_abc_answers = data.iloc[:, third:2*third].copy()
    binary_answers = data.iloc[:, 2*third:letter_index].copy()
    
    # Функция для назначения меток
    def assign_labels(indices, prefix):
        nonlocal question_counter
        for index in indices:
            question_labels[index] = f"{prefix}-{question_counter}"
            question_counter += 1
    
    # Извлечение и распознавание данных из оставшихся заголовков
    extra_headers = df.columns[letter_index:]
    parsed_headers = []
    binary_start_index = 11 + 2 * third

    question_labels = {}
    question_counter = 1
    current_index = binary_start_index + 1

    subject_question_info = {}
    category_scores = {}

    for header in extra_headers:
        lines = header.split('\n')
        if len(lines) == 4:
            subject = lines[0].strip()
            category = lines[1].strip()
            
            question_count_str = re.search(r'\d+(?:,\d{3})*(?:[.,]\d+)?', lines[2]).group() if re.search(r'\d+(?:,\d{3})*(?:[.,]\d+)?', lines[2]) else '0'
            score_per_question_str = re.search(r'\d+(?:,\d{3})*(?:[.,]\d+)?', lines[3]).group() if re.search(r'\d+(?:,\d{3})*(?:[.,]\d+)?', lines[3]) else '0'
            
            question_count = float(question_count_str.replace(',', '').replace('.', '').replace(',', '.')) if question_count_str.count(',') > 1 else float(question_count_str.replace(',', ''))
            score_per_question = float(score_per_question_str.replace(',', '').replace('.', '').replace(',', '.')) if score_per_question_str.count(',') > 1 else float(score_per_question_str.replace(',', ''))
            
            prefix = category[0].upper()
            indices = list(range(current_index, current_index + int(question_count)))
            assign_labels(indices, prefix)
            parsed_headers.append((subject, category, question_count, score_per_question, indices[0], indices[-1]))
            if subject not in subject_question_info:
                subject_question_info[subject] = {}
            subject_question_info[subject][category] = indices
            category_scores[(subject, category)] = score_per_question
            current_index += int(question_count)
    
    # Присвоение меток бинарным ответам
    binary_answers.columns = [question_labels.get(i, col) for i, col in enumerate(binary_answers.columns, start=binary_start_index + 1)]
    
    # Присвоение меток ABC ответам
    abc_answers.columns = binary_answers.columns.copy()

    # Создаем копию бинарных ответов для дальнейших вычислений
    numeric_df = binary_answers.copy()

    for subject, categories in subject_question_info.items():
        for category, indices in categories.items():
            original_col = f"{subject} {category} (Оригинал)"
            multiplied_col = f"{subject} {category} (Умноженный)"
            
            # Создание столбцов для оригинала и умноженного значений
            numeric_df[original_col] = numeric_df.loc[:, [question_labels[i] for i in indices]].sum(axis=1)
            numeric_df[multiplied_col] = numeric_df[original_col] * category_scores[(subject, category)]

    # Создание итоговых значений
    for subject, categories in subject_question_info.items():
        # Сумма оригиналов
        original_columns = [f"{subject} {category} (Оригинал)" for category in categories.keys()]
        numeric_df[f"{subject} Сумма (Оригинал)"] = numeric_df[original_columns].sum(axis=1)
        
        # Сумма умноженных значений
        multiplied_columns = [f"{subject} {category} (Умноженный)" for category in categories.keys()]
        numeric_df[f"{subject} Сумма (Умноженный)"] = numeric_df[multiplied_columns].sum(axis=1)
        
        # Процент
        numeric_df[f"{subject} Процент"] = (numeric_df[f"{subject} Сумма (Умноженный)"] / 40) * 100

    # Подсчет общего количества правильных ответов и общего балла
    numeric_df["Количество правильных ответов"] = numeric_df.filter(like="Сумма (Оригинал)").sum(axis=1)
    numeric_df["Общий балл"] = numeric_df.filter(like="Сумма (Умноженный)").sum(axis=1)
    
    # Рассчет общего процента
    num_subjects = len(subject_question_info)
    total_questions = num_subjects * 40
    numeric_df["Процент"] = (numeric_df["Общий балл"] / total_questions) * 100

    # Подсчет четных и нечетных ответов
    even_indices = [col for col in binary_answers.columns if int(re.search(r'\d+', col).group()) % 2 == 0]
    odd_indices = [col for col in binary_answers.columns if int(re.search(r'\d+', col).group()) % 2 != 0]

    numeric_df["Четные ответы"] = numeric_df[even_indices].sum(axis=1)
    numeric_df["Нечетные ответы"] = numeric_df[odd_indices].sum(axis=1)

    # Объединение обработанных данных
    processed_data = pd.concat([abc_answers, correct_abc_answers, numeric_df], axis=1)
    final_df = pd.concat([unchanged_data, processed_data], axis=1)

    # Сохранение результата в Excel
    final_df.to_excel(output_file_path, index=False)

    # Добавление новой операции: создание отдельных листов для каждого предмета и класса
    # Загрузка рабочего файла
    wb = load_workbook(output_file_path)

    # Получение информации о классах из столбца J (индекс 9)
    classes = df.iloc[:, 9].dropna().unique().astype(str)

    # Создание маппинга между метками вопросов и предметами
    question_label_to_subject = {}
    for subject, categories in subject_question_info.items():
        for category, indices in categories.items():
            for index in indices:
                q_label = question_labels.get(index)
                if q_label:
                    question_label_to_subject[q_label] = subject

    # Для каждого предмета и класса создаем листы и записываем данные
    for subject in subject_question_info.keys():
        # Получение соответствующих столбцов для предмета
        subject_cols = [col for col in binary_answers.columns if question_label_to_subject.get(col) == subject]
        
        for cls in classes:
            # Фильтрация данных по классу
            class_filter = df.iloc[:, 9].astype(str) == cls
            
            # Получение данных
            abc_data = abc_answers.loc[class_filter, subject_cols]
            binary_data = binary_answers.loc[class_filter, subject_cols]

            # Добавление первого столбца к каждой таблице без изменений
            first_column_data = df.loc[class_filter, df.columns[0]]
            abc_data.insert(0, 'PersonID', first_column_data.values)
            binary_data.insert(0, 'PersonID', first_column_data.values)
            
            # Создание листов ABC, БМ и Карта решаемости
            abc_sheet_name = f'Дистракторы_{cls}_{subject}'
            bm_sheet_name = f'БМ_{cls}_{subject}'
            karta_sheet_name = f'Карта решаемости_{cls}_{subject}'

            # Добавляем лист для ABC
            if abc_sheet_name not in wb.sheetnames:
                abc_sheet = wb.create_sheet(abc_sheet_name)
            else:
                abc_sheet = wb[abc_sheet_name]

            for r in dataframe_to_rows(abc_data, index=False, header=True):
                abc_sheet.append(r)

            # Добавляем лист для БМ
            if bm_sheet_name not in wb.sheetnames:
                bm_sheet = wb.create_sheet(bm_sheet_name)
            else:
                bm_sheet = wb[bm_sheet_name]

            for r in dataframe_to_rows(binary_data, index=False, header=True):
                bm_sheet.append(r)

            # Создаем лист для Карты решаемости (копирование данных из бинарного и их умножение на значения)
            if karta_sheet_name not in wb.sheetnames:
                karta_sheet = wb.create_sheet(karta_sheet_name)
            else:
                karta_sheet = wb[karta_sheet_name]

            multiplied_data = binary_data.copy()
            for category, indices in subject_question_info[subject].items():
                for col in [question_labels[i] for i in indices]:
                    multiplied_data[col] = multiplied_data[col] * category_scores[(subject, category)]
            
            for r in dataframe_to_rows(multiplied_data, index=False, header=True):
                karta_sheet.append(r)

    # Окрашивание столбцов
    colors = {
        "B-": "#1773c8",
        "Q-": "#ff7e25",
        "M-": "#ffc100",
        "Bilish": "#1773c8",
        "Qo‘llash": "#ff7e25",
        "Mulohaza": "#ffc100"
    }
    
    def color_columns(ws, headers, color_dict):
        for col in ws.columns:
            col_letter = col[0].column_letter
            header = col[0].value
            if header:
                header = str(header).strip()
                for key, color in color_dict.items():
                    if key in header:
                        fill = PatternFill(start_color=color.replace("#", ""), end_color=color.replace("#", ""), fill_type="solid")
                        for cell in col:
                            cell.fill = fill
                        break

    # Окрашивание столбцов для каждого листа
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        color_columns(ws, ws[1], colors)  # Окрашиваем заголовки первого ряда

    # Сохранение рабочего файла
    wb.save(output_file_path)

    print("Файл успешно обработан и сохранён:", output_file_path)

    # Open the saved file
    os.startfile(output_file_path)

    # Calculate execution time
    execution_time = time.time() - start_time
    print(f"Время выполнения: {execution_time:.2f} секунд")

    # Prevent terminal closure
    input("Нажмите Enter для выхода...")

process_excel_file()
